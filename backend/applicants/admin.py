from django.contrib import admin
from .models import Applicant
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


@admin.register(Applicant)
class ApplicantAdmin(admin.ModelAdmin):
    list_display = ['name', 'age', 'birth_date', 'email', 'phone_number', 'address', 'identification',
                    'name_guardian', 'email_guardian', 'phone_number_guardian', 'address_guardian', 'identification_guardian']
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Aspirantes"

        headers = [
            'Nombre Aspirante', 'Edad', 'Fecha de nacimiento',
            'Correo', 'Identificación', 'Teléfono', 'Motivación', 'Dirección',
            'Nombre del Acudiente', 'Correo del Acudiente', 'Teléfono del Acudiente',
            'Dirección del Acudiente', 'Identificación del Acudiente'
        ]
        sheet.append(headers)

        # Loop through selected applicants
        for obj in queryset:
            sheet.append([
                obj.name,
                obj.age,
                obj.birth_date,
                obj.email,
                obj.identification,
                obj.phone_number,
                obj.motivation,
                obj.address,
                obj.name_guardian,
                obj.email_guardian,
                obj.phone_number_guardian,
                obj.address_guardian,
                obj.identification_guardian
            ])

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[get_column_letter(
                column)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=aspirantes.xlsx'
        workbook.save(response)
        return response
    export_as_excel.short_description = "Exportar TODOS los apirantes a Excel"
